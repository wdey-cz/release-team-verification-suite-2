"""
TEMP: Convert an Excel sheet to customers.json.

Edit the two paths below, then run:
  python xlsx_to_customers_json.py

Requirements:
  python -m pip install openpyxl
"""

import json
from collections import OrderedDict
from pathlib import Path
from typing import Any, Dict, List, Tuple

import openpyxl


# ----------------------------
# EDIT THESE PATHS
# ----------------------------
INPUT_XLSX_PATH = Path(r"CustomerDB.xlsx")   # <-- change this
OUTPUT_JSON_PATH = Path(r"../assets/customers1.json")  # <-- change this
SHEET_NAME = None  # e.g. "Sheet1" or leave None to use the active sheet
# ----------------------------


EXPECTED_HEADERS = ["Customer Name", "Customer ID", "Role", "Username"]


def normalize_cell(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def find_header_row(ws) -> Tuple[int, Dict[str, int]]:
    """
    Find the row containing the expected headers.
    Returns:
      header_row_index_1_based, mapping header->column_index_0_based
    """
    for r in range(1, min(ws.max_row, 50) + 1):
        values = [normalize_cell(c.value) for c in ws[r]]
        if not any(values):
            continue

        mapping = {v: idx for idx, v in enumerate(values) if v}
        if all(h in mapping for h in EXPECTED_HEADERS):
            return r, {h: mapping[h] for h in EXPECTED_HEADERS}

    raise ValueError(f"Header row not found. Expected headers: {EXPECTED_HEADERS}")


def read_rows(ws, header_row: int, col_map: Dict[str, int]) -> List[Tuple[str, int, str, str]]:
    out: List[Tuple[str, int, str, str]] = []

    for r in range(header_row + 1, ws.max_row + 1):
        row_cells = ws[r]

        cust_name = normalize_cell(row_cells[col_map["Customer Name"]].value)
        cust_id_raw = row_cells[col_map["Customer ID"]].value
        role = normalize_cell(row_cells[col_map["Role"]].value)
        username = normalize_cell(row_cells[col_map["Username"]].value)

        # skip totally blank rows
        if not (cust_name or cust_id_raw or role or username):
            continue

        # basic validation
        if cust_name == "" or role == "" or username == "" or cust_id_raw in (None, ""):
            raise ValueError(
                f"Row {r} missing required fields: "
                f"Customer Name={cust_name!r}, Customer ID={cust_id_raw!r}, Role={role!r}, Username={username!r}"
            )

        try:
            cust_id = int(str(cust_id_raw).strip())
        except Exception as e:
            raise ValueError(f"Row {r} has non-integer Customer ID: {cust_id_raw!r}") from e

        out.append((cust_name, cust_id, role, username))

    return out


def build_customers_json(rows: List[Tuple[str, int, str, str]]) -> Dict[str, Any]:
    """
    Groups rows by customer_id.
    If duplicate (customer_id, role) appears, last one wins.
    """
    customers_by_id: Dict[int, Dict[str, Any]] = OrderedDict()

    for cust_name, cust_id, role, username in rows:
        if cust_id not in customers_by_id:
            customers_by_id[cust_id] = {
                "id": cust_id,
                "name": cust_name,
                "accounts": [],
                "_role_index": {},  # helper map for dedupe
            }

        cust = customers_by_id[cust_id]

        if cust["name"] != cust_name:
            print(
                f"Warning: Customer ID {cust_id} has multiple names: "
                f"{cust['name']!r} vs {cust_name!r}. Keeping {cust['name']!r}."
            )

        role_index: Dict[str, int] = cust["_role_index"]
        account = {"role": role, "username": username}

        if role in role_index:
            cust["accounts"][role_index[role]] = account
        else:
            role_index[role] = len(cust["accounts"])
            cust["accounts"].append(account)

    customers: List[Dict[str, Any]] = []
    for cust in customers_by_id.values():
        cust.pop("_role_index", None)
        customers.append(cust)

    return {"version": 1, "customers": customers}


def main():
    if not INPUT_XLSX_PATH.exists():
        raise FileNotFoundError(f"Input XLSX not found: {INPUT_XLSX_PATH}")

    wb = openpyxl.load_workbook(INPUT_XLSX_PATH, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME else wb.active

    header_row, col_map = find_header_row(ws)
    rows = read_rows(ws, header_row, col_map)
    payload = build_customers_json(rows)

    OUTPUT_JSON_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)

    print(f"Sheet: {ws.title}")
    print(f"Rows converted: {len(rows)}")
    print(f"Customers created: {len(payload['customers'])}")
    print(f"Wrote JSON to: {OUTPUT_JSON_PATH.resolve()}")


if __name__ == "__main__":
    main()
