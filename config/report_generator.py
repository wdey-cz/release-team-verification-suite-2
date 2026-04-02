from jinja2 import Environment, FileSystemLoader
from jinja2 import Environment, FileSystemLoader
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font
import re


class ReportGenerator:

    def __init__(self, db, log_fn, template_dir):
        self.db = db
        self.log = log_fn
        self.env = Environment(loader=FileSystemLoader(template_dir))

    def _split_csv(self, s):
        return [x.strip() for x in (s or "").split(",") if x.strip()]

    def _fetch_run_metadata(self, run_id):
        cursor = self.db.connection.cursor()

        cursor.execute("""
            SELECT browsers, clients, user_roles
            FROM test_runs
            WHERE run_id = ?;
        """, (run_id,))

        row = cursor.fetchone()
        if not row:
            raise ValueError(f"No data found for run_id={run_id}")

        return (
            self._split_csv(row[0]),
            self._split_csv(row[1]),
            self._split_csv(row[2])
        )

    def _fetch_logs(self, run_id, client, role, browser, only_types=None):
        cursor = self.db.connection.cursor()

        query = """
            SELECT timestamp, type, test_case_id, test_name,
                   message, status, time_taken_ms, comment, current_url
            FROM test_logs
            WHERE run_id = ?
              AND client_id = ?
              AND user_role = ?
              AND browser = ?
        """

        params = [run_id, client, role, browser]

        if only_types:
            query += " AND type IN ({})".format(",".join(["?"] * len(only_types)))
            params.extend(only_types)

        query += " ORDER BY id ASC"

        cursor.execute(query, params)
        return cursor.fetchall()

    def _transform_logs(self, logs):
        result = defaultdict(list)
        current_test = None

        for row in logs:
            ts, typ, tci, test_name, msg, status, timetaken, comment, url = row

            test_name = test_name or "NO_TEST_NAME"

            if typ == "test_case":
                current_test = {
                    "test_case_id": tci,
                    "test_name": test_name,
                    "test_case_name": msg,
                    "logs": [],
                    "status": status,
                    "time": timetaken,
                    "comment": comment,
                    "url": url
                }

                result[test_name].append(current_test)

            elif current_test:
                current_test["logs"].append({
                    "timestamp": ts,
                    "type": typ,
                    "message": msg,
                    "status": status,
                    "time": timetaken,
                    "comment": comment,
                    "url": url
                })

        return result

    # -------------------------
    # HTML REPORT
    # -------------------------
    def generate_html(self, run_id, output_dir, icon_path):
        browsers, clients, roles = self._fetch_run_metadata(run_id)

        data = {}

        for client in clients:
            for role in roles:
                for browser in browsers:
                    logs = self._fetch_logs(run_id, client, role, browser)
                    key = f"{client}_{role}_{browser}"
                    data[key] = self._transform_logs(logs)

        template = self.env.get_template("report_template.html")

        html = template.render(
            data=data,
            run_id=run_id,
            icon_path=icon_path
        )
        print(data)
        output_path = output_dir / f"Report_{run_id}.html"
        output_path.write_text(html, encoding="utf-8")

        self.log(f"[OK] HTML report generated at {output_path}")
        return output_path

    # -------------------------
    # XLSX REPORT
    # -------------------------
    def _sanitize_sheet_name(self, name):
        name = (name or "").strip() or "NO_TEST_NAME"
        name = re.sub(r"[:\\/?*\[\]]", "_", name)
        return name[:31]

    def _unique_sheet_title(self, wb, base):
        base = self._sanitize_sheet_name(base)
        if base not in wb.sheetnames:
            return base

        i = 2
        while True:
            suffix = f"_{i}"
            trimmed = base[:31 - len(suffix)]
            candidate = f"{trimmed}{suffix}"
            if candidate not in wb.sheetnames:
                return candidate
            i += 1

    def generate_xlsx(self, run_id, output_dir):
        browsers, clients, roles = self._fetch_run_metadata(run_id)

        header = ["Timestamp", "Type", "Test Case ID", "Test Name",
                  "Message", "Status", "Time Taken", "Comments", "Current URL"]

        header_font = Font(bold=True)

        for client in clients:
            for role in roles:
                for browser in browsers:

                    logs = self._fetch_logs(
                        run_id, client, role, browser,
                        only_types=["test_case", "heartbeat"]
                    )

                    xlsx_file = output_dir / f"report_{client}_{role}_{browser}.xlsx"

                    wb = Workbook()
                    wb.remove(wb.active)

                    sheets = {}

                    if not logs:
                        ws = wb.create_sheet("NO_DATA")
                        for c, h in enumerate(header, 1):
                            cell = ws.cell(row=1, column=c, value=h)
                            cell.font = header_font
                    else:
                        for row in logs:
                            ts, typ, tci, test_name, msg, status, timetaken, comment, url = row

                            tab_key = "HEARTBEAT" if typ == "heartbeat" else (test_name or "NO_TEST_NAME")

                            if tab_key not in sheets:
                                title = self._unique_sheet_title(wb, tab_key)
                                ws = wb.create_sheet(title)
                                sheets[tab_key] = ws

                                for c, h in enumerate(header, 1):
                                    cell = ws.cell(row=1, column=c, value=h)
                                    cell.font = header_font

                            sheets[tab_key].append(
                                [ts, typ, tci, test_name, msg, status, timetaken, comment, url]
                            )

                    wb.save(xlsx_file)

        self.log(f"[OK] XLSX reports generated at {output_dir}")
        return output_dir